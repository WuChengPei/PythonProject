import subprocess
import sys
import tkinter as tk
from tkinter import messagebox, ttk, Canvas, Entry, Button, PhotoImage
import time
from pathlib import Path
from tkinter import filedialog
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill
import pandas as pd
from copy import copy


def relative_to_assets(path: str) -> Path:
    if getattr(sys, 'frozen', False):  # PyInstaller 打包後
        base_path = Path(sys.executable).parent  # `.exe` 的所在目錄
    else:  # 開發環境
        base_path = Path(__file__).parent
    assets_path = base_path / "assets2"  # `assets` 應該在 `.exe` 同個資料夾
    full_path = assets_path / path

    if not full_path.exists():
        print(f"請確認 `assets` 資料夾是否正確！")
    return full_path


class CalculatorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("HIMR Calculator")
        self.root.geometry("554x301")
        self.root.configure(bg="#FFFFFF")

        icon_path = relative_to_assets("P.ico")  # 使用相對路徑函式
        if icon_path.exists():
            self.root.iconbitmap(str(icon_path))  # 設定 icon
        else:
            print("請確認 assets 目錄是否正確！")

        canvas = Canvas(
            root,
            bg="#FFFFFF",
            height=301,
            width=554,
            bd=0,
            highlightthickness=0,
            relief="ridge"
        )
        canvas.pack()
        # 先創建背景圖片，讓它在最底層
        self.image1 = PhotoImage(file=relative_to_assets("cc.png"))
        img_width = self.image1.width()
        img_height = self.image1.height()

        canvas_width = 554
        canvas_height = 301
        center_x = canvas_width / 2  # 277
        center_y = canvas_height / 2  # 150.5
        canvas.create_image(center_x, center_y, image=self.image1)

        # 儲存路徑的變量
        self.save_path = os.path.expanduser("~/Documents")  # 預設為文件夾

        # path button - 現在會顯示在背景圖片上方
        self.button_image_2 = PhotoImage(file=relative_to_assets("b2.png"))
        self.select_path_button = Button(
            canvas, 
            image=self.button_image_2, 
            text="Select the Path",                  
            compound="center",            
            font=("Cormorant SC Medium", 8, 'italic'),
            fg="black",
            borderwidth=0, 
            highlightthickness=0,
            bg="#F0EAE2", 
            activebackground="#FFFFFF", 
            command=self.select_save_path,
            relief="flat"
        )
        self.select_path_button.place(x=31.5, y=225, width=70, height=40)

        # try button - 現在會顯示在背景圖片上方
        self.button_image_3 = PhotoImage(file=relative_to_assets("b3.png")) 
        self.try_button = Button(
            canvas, 
            image=self.button_image_3, 
            text="Try →",                  
            compound="center",            
            font=("Cormorant SC Medium", 8, 'italic'),
            fg="black",
            borderwidth=0, 
            highlightthickness=0,
            bg="#B2AEA9", 
            activebackground="#FFFFFF", 
            command=self.select_save_path,
            relief="flat"
        )
        self.try_button.place(x=460, y=265, width=70, height=40)

        canvas_width = 554
        canvas_height = 301
        center_x = canvas_width / 2  # 277
        center_y = canvas_height / 2  # 150.5
        canvas.create_image(center_x, center_y, image=self.image1)

        

    def select_save_path(self):
        """選擇檔案儲存位置"""
        path = filedialog.askdirectory(initialdir=self.save_path)
        if path:
            self.save_path = path
            messagebox.showinfo("成功", f"已選擇儲存位置：{path}")

    def process_excel_data(self):
        """合併 Excel 檔案並處理數據，生成最終結果"""
        try:
            # 打開檔案選擇對話框，選擇包含 "查詢清單" 的 Excel 檔案
            query_file = filedialog.askopenfilename(
                title="選擇包含 '查詢清單' 的 Excel 檔案",
                filetypes=[("Excel files", "*.xlsx *.xls")]
            )
            if not query_file:
                return  # 用戶取消選擇

            # 載入用戶選擇的 Excel 檔案
            wb_user = load_workbook(query_file)
            if "查詢清單" not in wb_user.sheetnames:
                messagebox.showerror("錯誤", "選擇的檔案中不存在 '查詢清單' 工作表！")
                return
            ws_query = wb_user["查詢清單"]

            # 載入 assets2 中的 "113年健保申報量.xlsx"
            base_file = relative_to_assets("113年健保申報量.xlsx")
            wb_base = load_workbook(base_file)
            if "使用量" not in wb_base.sheetnames:
                messagebox.showerror("錯誤", "'113年健保申報量.xlsx' 中不存在 '使用量' 工作表！")
                return
            ws_source = wb_base["使用量"]

            # 如果目標檔案已有 "113年健保申報量" 工作表，則刪除
            if "113年健保申報量" in wb_user.sheetnames:
                del wb_user["113年健保申報量"]

            # 創建新工作表 "113年健保申報量"
            ws_target = wb_user.create_sheet("113年健保申報量")

            # 複製所有儲存格
            for row in ws_source.rows:
                for cell in row:
                    ws_target.cell(row=cell.row, column=cell.column, value=cell.value)

                    # 複製單元格樣式
                    if cell.has_style:
                        ws_target.cell(row=cell.row, column=cell.column).font = copy(cell.font)
                        ws_target.cell(row=cell.row, column=cell.column).border = copy(cell.border)
                        ws_target.cell(row=cell.row, column=cell.column).fill = copy(cell.fill)
                        ws_target.cell(row=cell.row, column=cell.column).number_format = cell.number_format
                        ws_target.cell(row=cell.row, column=cell.column).protection = copy(cell.protection)
                        ws_target.cell(row=cell.row, column=cell.column).alignment = copy(cell.alignment)

            # 複製列寬
            for col in ws_source.column_dimensions:
                ws_target.column_dimensions[col].width = ws_source.column_dimensions[col].width

            # 複製列高
            for row in ws_source.row_dimensions:
                ws_target.row_dimensions[row].height = ws_source.row_dimensions[row].height

            # 將合併後的檔案保存到臨時位置
            temp_file = os.path.join(self.save_path, "merged_temp.xlsx")
            wb_user.save(temp_file)

            # 載入臨時檔案中的 "113年健保申報量" 工作表
            wb_data = load_workbook(temp_file)
            ws_data = wb_data["113年健保申報量"]

            # 創建新的工作簿和工作表 "查詢結果"
            wb_output = Workbook()
            ws_output = wb_output.active
            ws_output.title = "查詢結果"

            # 取得查詢清單最後一行
            last_row_query = ws_query.max_row
            for row in range(last_row_query, 1, -1):
                if ws_query[f"B{row}"].value:
                    last_row_query = row
                    break
            if last_row_query < 2:
                messagebox.showerror("錯誤", "查詢清單 B 欄無資料！")
                return

            # 取得 "113年健保申報量" 資料範圍
            last_row_data = ws_data.max_row
            for row in range(last_row_data, 1, -1):
                if ws_data[f"H{row}"].value:
                    last_row_data = row
                    break
            last_col_data = ws_data.max_column
            for col in range(last_col_data, 1, -1):
                if ws_data.cell(1, col).value:
                    last_col_data = col
                    break

            # 設置標題列：A 欄為市佔率，其餘欄位右移
            ws_output.cell(1, 1).value = "市佔率"
            for col in range(1, last_col_data + 1):
                ws_output.cell(1, col + 1).value = ws_data.cell(1, col).value
            output_row = 2

            # 初始化字典，用於儲存健保分類分組名稱和藥品申報量總和
            dict_groups = {}

            # 迴圈處理每個藥品名稱
            for row in range(2, last_row_query + 1):
                drug_name = str(ws_query[f"B{row}"].value).strip() if ws_query[f"B{row}"].value else ""
                if not drug_name:
                    continue

                # 查找藥品對應的健保分類分組名稱
                found = False
                group_name = ""
                for i in range(2, last_row_data + 1):
                    if str(ws_data.cell(i, 4).value).strip() == drug_name:  # D 欄為藥品名稱
                        group_name = str(ws_data.cell(i, 8).value).strip()  # H 欄為健保分類分組名稱
                        found = True
                        break

                # 添加藥品名稱標題
                ws_output.cell(output_row, 1).value = f"藥品名稱: {drug_name}"
                ws_output.cell(output_row, 1).font = Font(bold=True)
                output_row += 1

                # 若未找到藥品或無健保分類分組名稱，輸出提示訊息
                if not found or not group_name:
                    ws_output.cell(output_row, 1).value = "未找到或無健保分類分組名稱！"
                    ws_output.cell(output_row, 1).font = Font(italic=True)
                    output_row += 1
                    output_row += 1  # 添加空白行
                    continue

                # 複製相同健保分類分組名稱的資料並記錄申報量
                for i in range(2, last_row_data + 1):
                    if str(ws_data.cell(i, 8).value).strip() == group_name:
                        # 複製整行資料（從 A 欄到最後一欄，右移到 B 欄）
                        for col in range(1, last_col_data + 1):
                            ws_output.cell(output_row, col + 1).value = ws_data.cell(i, col).value
                        # 記錄藥品申報量（L 欄，第 12 欄）
                        declaration_amount = ws_data.cell(i, 12).value
                        if isinstance(declaration_amount, (int, float)):
                            if group_name not in dict_groups:
                                dict_groups[group_name] = declaration_amount
                            else:
                                dict_groups[group_name] += declaration_amount
                        output_row += 1

                # 添加空白行以分隔不同藥品
                output_row += 1

            # 計算並寫入市佔率
            last_row_output = ws_output.max_row
            for row in range(2, last_row_output + 1):
                # 跳過標題行和空白行
                cell_a_value = str(ws_output.cell(row, 1).value or "")
                cell_b_value = ws_output.cell(row, 2).value or ""
                if cell_a_value.startswith("藥品名稱: ") or not cell_b_value:
                    continue
                group_name = str(ws_output.cell(row, 9).value).strip()  # H 欄右移到 I 欄（第 9 欄）
                declaration_amount = ws_output.cell(row, 13).value  # L 欄右移到 M 欄（第 13 欄）
                if group_name in dict_groups and isinstance(declaration_amount, (int, float)):
                    total_amount = dict_groups.get(group_name, 0)
                    if total_amount > 0:
                        market_share = declaration_amount / total_amount
                        ws_output.cell(row, 1).value = market_share
                        ws_output.cell(row, 1).number_format = "0.00%"  # 格式化為百分比
                    else:
                        ws_output.cell(row, 1).value = 0

            # 高亮顯示查詢清單中的藥品名稱
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            for row in range(2, last_row_query + 1):
                drug_name = str(ws_query[f"B{row}"].value).strip() if ws_query[f"B{row}"].value else ""
                if not drug_name:
                    continue
                for i in range(2, last_row_output + 1):
                    cell_a_value = str(ws_output.cell(i, 1).value or "")
                    cell_e_value = ws_output.cell(i, 5).value or ""  # D 欄右移到 E 欄（第 5 欄）
                    if cell_a_value.startswith("藥品名稱: ") or not cell_e_value:
                        continue
                    if str(cell_e_value).strip() == drug_name:
                        for col in range(1, last_col_data + 2):
                            ws_output.cell(i, col).fill = yellow_fill

            # 儲存結果
            output_file = os.path.join(self.save_path, "Results.xlsx")
            wb_output.save(output_file)

            # 自動調整欄寬（openpyxl 不支援直接 AutoFit，手動設置近似值）
            for col in ws_output.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws_output.column_dimensions[column].width = adjusted_width

            # 儲存最終結果
            wb_output.save(output_file)

            # 刪除臨時合併檔案
            try:
                os.remove(temp_file)
            except:
                pass

            # 顯示成功訊息
            messagebox.showinfo("Success", "Easy Peasy ><")

        except Exception as e:
            # 確保臨時檔案被刪除（即使發生錯誤）
            try:
                if 'temp_file' in locals() and os.path.exists(temp_file):
                    os.remove(temp_file)
            except:
                pass
            messagebox.showerror("錯誤", f"處理檔案時發生錯誤：{str(e)}")


if __name__ == "__main__":
    root = tk.Tk()
    app = CalculatorApp(root)
    root.resizable(False, False)
    root.mainloop()
